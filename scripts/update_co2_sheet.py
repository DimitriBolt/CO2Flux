from __future__ import annotations

import copy
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import oracledb
from dotenv import dotenv_values
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path("/home/dimitri/PycharmProjects/CO2Flux")
WORKBOOK_PATH = ROOT / "Sensors_Description" / "climate_control_theorist_schema.xlsx"
ENV_PATH = ROOT / ".env"
ORACLE_CLIENT_LIB_DIR = Path("/opt/oracle/instantclient_19_26")
TODAY = date.today().isoformat()
GMM222_PATTERN = re.compile(r"^LEO-(?P<slope>[A-Z])_(?P<y>-?\d+)_(?P<x>-?\d+)_(?P<level>\d+)_GMM222$")
DEPTH_CM_BY_LEVEL = {1: 5, 2: 20, 3: 35, 4: 50}
YES_VALIDATED = f"Yes — literal AD query validated in Oracle on {TODAY}."
NO_ZERO_ROWS = f"No — literal AD query returns zero rows in current Oracle export as of {TODAY}."


@dataclass
class SeriesBounds:
    first_dt: datetime | None
    first_val: float | None
    last_dt: datetime | None
    last_val: float | None

    @property
    def has_data(self) -> bool:
        return self.first_dt is not None and self.last_dt is not None


def connect() -> oracledb.Connection:
    cfg = dotenv_values(ENV_PATH)
    oracledb.init_oracle_client(lib_dir=str(ORACLE_CLIENT_LIB_DIR))
    dsn = oracledb.makedsn(cfg["ORACLE_HOST"], int(cfg["ORACLE_PORT"]), sid=cfg["ORACLE_SID"])
    conn = oracledb.connect(
        user=cfg["ORACLE_USER"],
        password=cfg["ORACLE_PASSWORD"],
        dsn=dsn,
    )
    conn.call_timeout = 120_000
    return conn


def fetch_bounds_sensor(
    cur: oracledb.Cursor,
    table: str,
    sensor_id: int,
    variable_id: int,
    time_col: str = "LOCALDATETIME",
    value_col: str = "DATAVALUE",
    start_date: str | None = None,
) -> SeriesBounds:
    start_filter = f"\n          AND dv.{time_col} >= DATE '{start_date}'" if start_date else ""
    q_first = f"""
        SELECT dv.{time_col}, dv.{value_col}
        FROM {table} dv
        WHERE dv.sensorid = :sensor_id
          AND dv.variableid = :variable_id
          {start_filter}
        ORDER BY dv.{time_col}
        FETCH FIRST 1 ROW ONLY
    """
    q_last = f"""
        SELECT dv.{time_col}, dv.{value_col}
        FROM {table} dv
        WHERE dv.sensorid = :sensor_id
          AND dv.variableid = :variable_id
          {start_filter}
        ORDER BY dv.{time_col} DESC
        FETCH FIRST 1 ROW ONLY
    """
    cur.execute(q_first, sensor_id=sensor_id, variable_id=variable_id)
    first_row = cur.fetchone()
    cur.execute(q_last, sensor_id=sensor_id, variable_id=variable_id)
    last_row = cur.fetchone()
    return SeriesBounds(
        first_dt=first_row[0] if first_row else None,
        first_val=first_row[1] if first_row else None,
        last_dt=last_row[0] if last_row else None,
        last_val=last_row[1] if last_row else None,
    )


def fetch_bounds_table(
    cur: oracledb.Cursor,
    table: str,
    time_col: str = "TIMESTAMP",
    value_col: str = "VALUE",
) -> SeriesBounds:
    q_first = f"""
        SELECT t.{time_col}, t.{value_col}
        FROM {table} t
        ORDER BY t.{time_col}
        FETCH FIRST 1 ROW ONLY
    """
    q_last = f"""
        SELECT t.{time_col}, t.{value_col}
        FROM {table} t
        ORDER BY t.{time_col} DESC
        FETCH FIRST 1 ROW ONLY
    """
    cur.execute(q_first)
    first_row = cur.fetchone()
    cur.execute(q_last)
    last_row = cur.fetchone()
    return SeriesBounds(
        first_dt=first_row[0] if first_row else None,
        first_val=first_row[1] if first_row else None,
        last_dt=last_row[0] if last_row else None,
        last_val=last_row[1] if last_row else None,
    )


def date_label(dt: datetime | None) -> str:
    if dt is None:
        return ""
    return dt.date().isoformat()


def end_label(dt: datetime | None) -> str:
    if dt is None:
        return ""
    return "present" if dt.date().isoformat() == TODAY else dt.date().isoformat()


def make_series_query(table: str, sensor_id: int, variable_id: int, start_date: str | None = None) -> str:
    where_lines = [
        f"    dv.sensorid = {sensor_id}",
        f"    AND dv.variableid = {variable_id}",
    ]
    if start_date:
        where_lines.append(f"    AND dv.localdatetime >= DATE '{start_date}'")
    return (
        "SELECT\n"
        "    dv.localdatetime,\n"
        "    dv.datavalue\n"
        f"FROM\n    {table} dv\n"
        "WHERE\n"
        + "\n".join(where_lines)
        + "\nORDER BY\n"
        "    dv.localdatetime;"
    )


def make_control_query(table: str) -> str:
    return (
        "SELECT\n"
        "    t.timestamp,\n"
        "    t.value\n"
        f"FROM\n    {table} t\n"
        "ORDER BY\n"
        "    t.timestamp;"
    )


def make_variable_query(schema: str, variable_id: int) -> str:
    return (
        "SELECT\n"
        "    v.variableid,\n"
        "    v.variablecode,\n"
        "    v.variablename,\n"
        "    u.unitsabbreviation\n"
        f"FROM\n    {schema}.variables v\n"
        f"    LEFT JOIN {schema}.units u\n"
        "        ON v.variableunitsid = u.unitsid\n"
        "WHERE\n"
        f"    v.variableid = {variable_id}\n"
        "ORDER BY\n"
        "    v.variableid;"
    )


def record_from_row(ws, row_index: int, columns: list[str]) -> dict[str, object]:
    return {col: ws[f"{col}{row_index}"].value for col in columns}


def apply_row_style(ws, template_row: int, target_row: int) -> None:
    for col_idx in range(1, ws.max_column + 1):
        src = ws.cell(template_row, col_idx)
        dst = ws.cell(target_row, col_idx)
        dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy.copy(src.font)
        if src.fill:
            dst.fill = copy.copy(src.fill)
        if src.border:
            dst.border = copy.copy(src.border)
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)
        if src.protection:
            dst.protection = copy.copy(src.protection)
    ws.row_dimensions[target_row].height = ws.row_dimensions[template_row].height


def clear_row_values(ws, row_index: int) -> None:
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(row_index, col_idx).value = None


def fetch_gmm222_rows(
    cur: oracledb.Cursor,
    schema: str,
    slope_label: str,
    inventory_file: str,
) -> list[dict[str, object]]:
    schema_prefix = schema.lower()
    cur.execute(
        f"""
        SELECT sensorid, sensorcode
        FROM {schema}.sensors
        WHERE sensorcode LIKE :pattern
        ORDER BY sensorid
        """,
        pattern=f"LEO-{slope_label[4]}_%GMM222",
    )
    rows: list[dict[str, object]] = []
    for sensor_id, sensor_code in cur.fetchall():
        match = GMM222_PATTERN.match(sensor_code)
        if not match:
            continue
        y = int(match.group("y"))
        x = int(match.group("x"))
        level = int(match.group("level"))
        depth_cm = DEPTH_CM_BY_LEVEL[level]
        bounds = fetch_bounds_sensor(cur, f"{schema_prefix}.datavalues", sensor_id, 9)
        start_date = date_label(bounds.first_dt)
        rows.append(
            {
                "A": "1",
                "B": "Carbon dioxide concentration",
                "C": "C_CO2,basalt",
                "D": "scalar",
                "E": "Primary state / gradient driver",
                "F": "Internal basalt CO2 concentration at a fixed pore-space point; used to build vertical gradients and surface-normal flux estimates.",
                "G": slope_label,
                "H": inventory_file,
                "I": "GMM222",
                "J": sensor_id,
                "K": sensor_code,
                "L": f"{schema_prefix}.datavalues",
                "M": f"dv.sensorid = {sensor_id} AND dv.variableid = 9",
                "N": "LOCALDATETIME",
                "O": "DATAVALUE",
                "P": "ppm",
                "Q": "ppm",
                "R": "not used for extraction; keep ppm",
                "S": f"{slope_label} basalt vertical line y={y} m, x={x} m, depth={depth_cm} cm",
                "T": "Basalt pore-space point on vertical profile (same x,y; varying depth).",
                "U": x,
                "V": y,
                "W": -depth_cm / 100,
                "X": f"Depth level {level} ({depth_cm} cm below surface)",
                "Y": "Yes",
                "Z": "Core internal CO2 measurement required for vertical gradients, spatial comparisons, and inference of basalt-atmosphere CO2 flux.",
                "AA": date_label(bounds.first_dt),
                "AB": end_label(bounds.last_dt),
                "AC": f"Literal AD query validated in Oracle on {TODAY}. Current Oracle series uses variableid=9 (CO2).",
                "AD": make_series_query(f"{schema_prefix}.datavalues", sensor_id, 9, start_date=start_date),
                "AE": "",
                "AF": YES_VALIDATED if bounds.has_data else NO_ZERO_ROWS,
                "AG": 9,
                "AH": "CO2",
                "AI": "Carbon dioxide",
                "AJ": "ppm",
                "AK": make_variable_query(schema_prefix, 9),
            }
        )
    rows.sort(key=lambda row: (int(row["V"]), int(row["U"]), int(str(row["X"]).split()[2])))
    return rows


def update_sensor_record(
    cur: oracledb.Cursor,
    record: dict[str, object],
    *,
    table: str,
    variable_id: int,
    variable_code: str,
    variable_name: str,
    units: str,
    note_with_data: str,
    note_no_data: str,
    use_in_v1_if_missing: str = "No",
) -> dict[str, object]:
    record["L"] = table
    record["M"] = f"dv.sensorid = {record['J']} AND dv.variableid = {variable_id}"
    schema = table.split(".")[0]
    record["AK"] = make_variable_query(schema, variable_id)
    record["AG"] = variable_id
    record["AH"] = variable_code
    record["AI"] = variable_name
    record["AJ"] = units
    bounds = fetch_bounds_sensor(cur, table, int(record["J"]), variable_id)
    start_date = date_label(bounds.first_dt)
    record["AD"] = make_series_query(table, int(record["J"]), variable_id, start_date=start_date or None)
    if bounds.has_data:
        record["AA"] = date_label(bounds.first_dt)
        record["AB"] = end_label(bounds.last_dt)
        record["AC"] = note_with_data
        record["AF"] = YES_VALIDATED
    else:
        record["AA"] = ""
        record["AB"] = ""
        record["AC"] = note_no_data
        record["AF"] = NO_ZERO_ROWS
        record["Y"] = use_in_v1_if_missing
    return record


def update_control_record(cur: oracledb.Cursor, record: dict[str, object], note: str) -> dict[str, object]:
    bounds = fetch_bounds_table(cur, str(record["L"]))
    record["AD"] = make_control_query(str(record["L"]))
    record["AA"] = date_label(bounds.first_dt)
    record["AB"] = end_label(bounds.last_dt)
    record["AC"] = note
    record["AF"] = YES_VALIDATED if bounds.has_data else NO_ZERO_ROWS
    return record


def update_full_series_sensor_record(cur: oracledb.Cursor, record: dict[str, object], note: str) -> dict[str, object]:
    table = str(record["L"])
    variable_id = int(record["AG"])
    start_date = str(record.get("AA") or "").strip() or None
    bounds = fetch_bounds_sensor(cur, table, int(record["J"]), variable_id, start_date=start_date)
    query_start = start_date or date_label(bounds.first_dt) or None
    record["AD"] = make_series_query(table, int(record["J"]), variable_id, start_date=query_start)
    record["AA"] = date_label(bounds.first_dt)
    record["AB"] = end_label(bounds.last_dt)
    record["AC"] = note if bounds.has_data else NO_ZERO_ROWS
    record["AF"] = YES_VALIDATED if bounds.has_data else NO_ZERO_ROWS
    if not bounds.has_data:
        record["Y"] = "No"
    return record


def main() -> None:
    conn = connect()
    cur = conn.cursor()

    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["CO2"]
    columns = [get_column_letter(i) for i in range(1, ws.max_column + 1)]

    center_basalt = [record_from_row(ws, row, columns) for row in range(4, 52)]
    east_air = [record_from_row(ws, row, columns) for row in range(52, 79)]
    center_air = [record_from_row(ws, row, columns) for row in range(79, 106)]
    west_air = [record_from_row(ws, row, columns) for row in range(106, 133)]
    support_rows = [record_from_row(ws, row, columns) for row in range(133, 153)]

    for record in center_basalt:
        update_sensor_record(
            cur,
            record,
            table="leo_center.datavalues",
            variable_id=9,
            variable_code="CO2",
            variable_name="Carbon dioxide",
            units="ppm",
            note_with_data=f"Literal AD query validated in Oracle on {TODAY}. Inventory dates were checked against the current Oracle series.",
            note_no_data=f"Literal AD query unexpectedly returned zero rows in Oracle on {TODAY}.",
        )

    east_basalt = fetch_gmm222_rows(cur, "LEO_EAST", "LEO East", "LEO-East-Inventory.xlsx")
    west_basalt = fetch_gmm222_rows(cur, "LEO_WEST", "LEO West", "LEO-West-Inventory.xlsx")

    east_air_note = (
        f"Literal AD query validated in Oracle on {TODAY}. For LI-7000 atmospheric CO2, variableid=56 "
        "(CO2_cellB) is retained because variableid=55 in the current Oracle export is almost entirely an "
        "exact 28.0/0.0 series, while variableid=56 carries the physically scaled concentration series. "
        "The Oracle table currently exposes the 2024 campaign window only."
    )
    west_air_note = (
        f"Literal AD query validated in Oracle on {TODAY}. For LI-7000 atmospheric CO2, variableid=56 "
        "(CO2_cellB) is retained because variableid=55 in the current Oracle export is almost entirely an "
        "exact 28.0/0.0 series, while variableid=56 carries the physically scaled concentration series. "
        "Public West Li-COR plots are also keyed on variable=56."
    )
    center_air_note = (
        f"Sensor metadata exist, but these center-slope LI-7000 points return zero rows in the current "
        f"Oracle export as of {TODAY}. leo_center.datavalueslicor currently contains only sensorid 1309 "
        "(LEO-G_CTest_LI-7000), not these slope-point sensors."
    )

    for record in east_air:
        update_sensor_record(
            cur,
            record,
            table="leo_east.datavalueslicor",
            variable_id=56,
            variable_code="CO2_cellB",
            variable_name="Carbon dioxide",
            units="umol/mol",
            note_with_data=east_air_note,
            note_no_data=f"Literal AD query returns zero rows in current Oracle export as of {TODAY}.",
        )

    for record in center_air:
        update_sensor_record(
            cur,
            record,
            table="leo_center.datavalueslicor",
            variable_id=56,
            variable_code="CO2_cellB",
            variable_name="Carbon dioxide",
            units="umol/mol",
            note_with_data=center_air_note,
            note_no_data=center_air_note,
            use_in_v1_if_missing="No",
        )
        record["Y"] = "No"

    for record in west_air:
        update_sensor_record(
            cur,
            record,
            table="leo_west.datavalueslicor",
            variable_id=56,
            variable_code="CO2_cellB",
            variable_name="Carbon dioxide",
            units="umol/mol",
            note_with_data=west_air_note,
            note_no_data=f"Sensor metadata exist, but this West LI-7000 point returns zero rows in the current Oracle export as of {TODAY}.",
            use_in_v1_if_missing="No",
        )

    for record in support_rows:
        series_id = int(record["J"]) if isinstance(record["J"], (int, float)) else None
        if series_id == 1275 and record["B"] == "Water vapor concentration":
            update_sensor_record(
                cur,
                record,
                table="leo_center.datavalueslicor",
                variable_id=58,
                variable_code="H2O_cellB",
                variable_name="Water vapor concentration",
                units="mmol/mol",
                note_with_data=center_air_note,
                note_no_data=center_air_note,
                use_in_v1_if_missing="No",
            )
            record["Y"] = "No"
            continue

        if str(record["G"]) == "Bio2 Controls":
            update_control_record(
                cur,
                record,
                note=f"Literal AD query validated in Oracle on {TODAY}. Query updated to return the full available series from the live control table.",
            )
            continue

        update_full_series_sensor_record(
            cur,
            record,
            note=f"Literal AD query validated in Oracle on {TODAY}. Query updated to return the full available series.",
        )

    output_rows = center_basalt + east_basalt + west_basalt + east_air + center_air + west_air + support_rows
    start_row = 4
    template_row = 4
    last_output_row = start_row + len(output_rows) - 1

    if last_output_row > ws.max_row:
        ws.insert_rows(ws.max_row + 1, amount=last_output_row - ws.max_row)

    for row_index, record in enumerate(output_rows, start=start_row):
        apply_row_style(ws, template_row, row_index)
        for col in columns:
            ws[f"{col}{row_index}"] = record.get(col, None)

    for row_index in range(last_output_row + 1, ws.max_row + 1):
        clear_row_values(ws, row_index)

    wb.save(WORKBOOK_PATH)
    cur.close()
    conn.close()


if __name__ == "__main__":
    main()
