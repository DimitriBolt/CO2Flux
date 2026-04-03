from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from air_co2_series import AirCO2Series, DEFAULT_AIR_HEIGHT_M


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_WORKBOOK_PATH = PROJECT_ROOT / "Sensors_Description" / "variables_schema.xlsx"


def _float_matches(left: float | int | None, right: float | int) -> bool:
    if left is None:
        return False
    return abs(float(left) - float(right)) < 1e-9


@dataclass(slots=True)
class AirCO2Catalog:
    workbook_path: Path = DEFAULT_WORKBOOK_PATH
    _sensors: list[AirCO2Series] = field(default_factory=list, init=False, repr=False)

    def __post_init__(self) -> None:
        self.workbook_path = Path(self.workbook_path)
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")
        self._sensors = self._load_sensors()

    def _load_sensors(self) -> list[AirCO2Series]:
        workbook = load_workbook(self.workbook_path, data_only=True)
        try:
            sheet = workbook["CO2"]
            sensors: list[AirCO2Series] = []
            for row_idx in range(4, sheet.max_row + 1):
                sensor = self.from_workbook_row(sheet, row_idx)
                if sensor is not None:
                    sensors.append(sensor)
            return sensors
        finally:
            workbook.close()

    def from_workbook_row(self, sheet: Worksheet, row_idx: int) -> AirCO2Series | None:
        if sheet[f"B{row_idx}"].value != "C_CO2,air":
            return None
        if sheet[f"K{row_idx}"].value != "LI-COR":
            return None

        sensor_id = sheet[f"L{row_idx}"].value
        sensor_code = sheet[f"M{row_idx}"].value
        table_name = sheet[f"N{row_idx}"].value
        variable_id = sheet[f"AD{row_idx}"].value
        slope = sheet[f"I{row_idx}"].value
        x_coord_m = sheet[f"C{row_idx}"].value
        y_coord_m = sheet[f"D{row_idx}"].value
        height_m = sheet[f"E{row_idx}"].value
        units = sheet[f"AG{row_idx}"].value or "ppm"

        required_values = {
            "sensor_id": sensor_id,
            "table_name": table_name,
            "variable_id": variable_id,
            "slope": slope,
            "x_coord_m": x_coord_m,
            "y_coord_m": y_coord_m,
            "height_m": height_m,
        }
        if any(value is None for value in required_values.values()):
            return None

        return AirCO2Series(
            table_name=str(table_name),
            sensor_id=int(sensor_id),
            sensor_code=str(sensor_code) if sensor_code is not None else None,
            variable_id=int(variable_id),
            slope=str(slope),
            x_coord_m=float(x_coord_m),
            y_coord_m=float(y_coord_m),
            height_m=float(height_m),
            units=str(units),
        )

    def list_sensors(self, *, slope: str | None = None) -> list[AirCO2Series]:
        if slope is None:
            return list(self._sensors)
        return [sensor for sensor in self._sensors if sensor.slope == slope]

    def find_sensor(
        self,
        slope: str,
        x_coord_m: float,
        y_coord_m: float,
        *,
        height_m: float = DEFAULT_AIR_HEIGHT_M,
    ) -> AirCO2Series:
        matches = [
            sensor
            for sensor in self._sensors
            if sensor.slope == slope
            and _float_matches(sensor.x_coord_m, x_coord_m)
            and _float_matches(sensor.y_coord_m, y_coord_m)
            and _float_matches(sensor.height_m, height_m)
        ]

        if not matches:
            raise ValueError(
                "No atmospheric CO2 row was found for "
                f"slope={slope!r}, X={x_coord_m}, Y={y_coord_m}, Z={height_m}."
            )
        if len(matches) > 1:
            sensor_codes = ", ".join(
                sensor.sensor_code or f"sensorid={sensor.sensor_id}" for sensor in matches
            )
            raise ValueError(
                "Multiple atmospheric CO2 rows matched the requested surface-air point: "
                f"{sensor_codes}"
            )
        return matches[0]

    def get_sensor(
        self,
        slope: str,
        x_coord_m: float,
        y_coord_m: float,
        *,
        height_m: float = DEFAULT_AIR_HEIGHT_M,
    ) -> AirCO2Series:
        return self.find_sensor(
            slope=slope,
            x_coord_m=x_coord_m,
            y_coord_m=y_coord_m,
            height_m=height_m,
        )


__all__ = [
    "AirCO2Catalog",
    "DEFAULT_WORKBOOK_PATH",
]
