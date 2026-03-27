from __future__ import annotations

import os
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path


DEFAULT_ORACLE_CLIENT_LIB_DIR = Path("/opt/oracle/instantclient_19_26")
LOCAL_LIB_DIR = Path.home() / ".local/lib"
ORACLE_ENV_READY_FLAG = "CO2_PROFILE_ORACLE_ENV_READY"


def _ensure_oracle_runtime_env() -> None:
    oracle_client_lib_dir = Path(
        os.getenv("ORACLE_CLIENT_LIB_DIR", str(DEFAULT_ORACLE_CLIENT_LIB_DIR))
    ).expanduser()
    if not oracle_client_lib_dir.exists():
        return

    required_entries = [str(oracle_client_lib_dir)]
    if LOCAL_LIB_DIR.exists():
        required_entries.append(str(LOCAL_LIB_DIR))

    current_entries = [
        entry for entry in os.getenv("LD_LIBRARY_PATH", "").split(os.pathsep) if entry
    ]
    missing_entries = [entry for entry in required_entries if entry not in current_entries]

    os.environ.setdefault("ORACLE_CLIENT_LIB_DIR", str(oracle_client_lib_dir))
    if not missing_entries or os.getenv(ORACLE_ENV_READY_FLAG) == "1":
        return

    new_env = os.environ.copy()
    new_env["ORACLE_CLIENT_LIB_DIR"] = str(oracle_client_lib_dir)
    new_env["LD_LIBRARY_PATH"] = os.pathsep.join(required_entries + current_entries)
    new_env[ORACLE_ENV_READY_FLAG] = "1"
    os.execve(sys.executable, [sys.executable, *sys.argv], new_env)


_ensure_oracle_runtime_env()

try:
    import matplotlib

    matplotlib.use("Agg")

    import matplotlib.pyplot as plt
    from matplotlib.animation import FuncAnimation, PillowWriter
    from openpyxl import load_workbook
    import oracledb
    from dotenv import dotenv_values
except Exception as exc:  # pragma: no cover - startup guard
    raise RuntimeError(
        "This script expects a Python interpreter where matplotlib, openpyxl, "
        "oracledb, Pillow, and python-dotenv are available. In this project, "
        "the known working choice is the default `python3` / Python 3.13 "
        "interpreter rather than `/usr/bin/python3.14`."
    ) from exc


# Editable parameters
# Allowed SLOPE values:
#   "LEO Center"
#   "LEO East"
#   "LEO West"
SLOPE = "LEO West"
X_COORD_M = -4  # Allowed X values [m]: -4, -1, 1, 4
Y_COORD_M = +10   # Allowed Y values [m]: 4, 10, 18, 24

# Date format:
#   "YYYY-Mon-DD HH:MM"  -> example: "2026-Mar-25 00:00"
#   "YYYY-Mon-DD"        -> start uses 00:00, end uses 23:59
#   END_DATE = ""        -> run until the last available observation
START_DATE = "2026-Mar-25 00:00"
END_DATE =   "2026-Mar-25 23:59"

# Shared plotting limits for all graphs
CO2_AXIS_MIN_PPM = 0
CO2_AXIS_MAX_PPM = 8000

# Rendering parameters
FRAME_DURATION_MS = 250
BAR_HEIGHT_M = 0.035
BAR_COLOR = "#2E5EAA"
SURFACE_LINE_COLOR = "#222222"
GRID_COLOR = "#CCCCCC"


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
WORKBOOK_PATH = SCRIPT_DIR / "climate_control_theorist_schema.xlsx"
ENV_PATH = PROJECT_ROOT / ".env"
OUTPUT_DIR = SCRIPT_DIR

DISPLAY_DEPTHS_M = [0.0, -0.05, -0.20, -0.35, -0.50]
NO_DATA_VALUE = -9999.0
DATETIME_WITH_TIME_FMT = "%Y-%b-%d %H:%M"
DATE_ONLY_FMT = "%Y-%b-%d"


@dataclass(frozen=True)
class ProfileSensor:
    sensor_id: int
    sensor_code: str
    table_name: str
    variable_id: int
    depth_m: float


@dataclass(frozen=True)
class Measurement:
    timestamp: datetime
    value: float


class SinglePlayPillowWriter(PillowWriter):
    """Save GIF without the Netscape loop extension so it plays once."""

    def finish(self) -> None:
        self._frames[0].save(
            self.outfile,
            save_all=True,
            append_images=self._frames[1:],
            duration=int(1000 / self.fps),
        )


def parse_user_datetime(raw_value: str, *, is_end: bool) -> datetime | None:
    text = raw_value.strip()
    if not text:
        return None

    for fmt in (DATETIME_WITH_TIME_FMT, DATE_ONLY_FMT):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt == DATE_ONLY_FMT and is_end:
                return parsed.replace(hour=23, minute=59)
            return parsed
        except ValueError:
            continue

    raise ValueError(
        f"Unsupported date format: {raw_value!r}. "
        f"Use '{DATETIME_WITH_TIME_FMT}' or '{DATE_ONLY_FMT}'."
    )


def connect_to_oracle() -> oracledb.Connection:
    cfg = dotenv_values(ENV_PATH)

    # Try to initialize Thick mode by discovering Oracle Instant Client.
    candidates: list[Path] = []

    # 1) Explicit environment variable
    env_lib_dir = os.getenv("ORACLE_CLIENT_LIB_DIR")
    if env_lib_dir:
        p = Path(env_lib_dir).expanduser()
        if p.exists():
            candidates.append(p)

    # 1b) .env configuration fallback
    cfg_lib_dir = cfg.get("ORACLE_CLIENT_LIB_DIR")
    if cfg_lib_dir:
        p = Path(str(cfg_lib_dir)).expanduser()
        if p.exists():
            candidates.append(p)

    # 2) Project default (Linux path, harmless on Windows if it doesn't exist)
    if DEFAULT_ORACLE_CLIENT_LIB_DIR.exists():
        candidates.append(DEFAULT_ORACLE_CLIENT_LIB_DIR)

    # 3) Platform-specific check helpers
    def _has_client_lib(path: Path) -> bool:
        if os.name == "nt":
            return (path / "oci.dll").exists()
        elif sys.platform == "darwin":
            return (path / "libclntsh.dylib").exists()
        else:
            return (path / "libclntsh.so").exists()

    # 4) Windows-specific discovery
    if os.name == "nt":
        typical_bases = [
            Path(r"C:\oracle"),
            Path(r"C:\oracle\instantclient"),
            Path(r"C:\instantclient"),
            Path(r"C:\Program Files\Oracle"),
            Path.home() / "oracle",
        ]
        for base in typical_bases:
            if not base.exists():
                continue
            if _has_client_lib(base):
                candidates.append(base)
            try:
                for sub in base.glob("**/instantclient*"):
                    if sub.is_dir():
                        candidates.append(sub)
            except Exception:
                pass

        # Also scan PATH entries for a folder that has client library
        for entry in os.environ.get("PATH", "").split(os.pathsep):
            if not entry:
                continue
            p = Path(entry)
            if p.exists() and _has_client_lib(p):
                candidates.append(p)

    # Deduplicate keeping order
    seen: set[str] = set()
    unique_candidates: list[str] = []
    for p in candidates:
        s = str(p)
        if s not in seen:
            seen.add(s)
            unique_candidates.append(s)

    used_lib_dir: str | None = None
    for lib_dir in unique_candidates:
        try:
            oracledb.init_oracle_client(lib_dir=lib_dir)
            used_lib_dir = lib_dir
            break
        except Exception as init_exc:
            print(
                f"Warning: failed to initialize Oracle Client at {lib_dir!r}: {init_exc}",
                file=sys.stderr,
            )

    # 5) If not found on Windows, ask interactively before attempting any connection in thin mode
    if used_lib_dir is None and os.name == "nt" and sys.stdin and sys.stdin.isatty():
        try:
            user_input = input(
                "Oracle Instant Client required for encryption. "
                "Enter full path to the folder containing oci.dll (or leave empty to skip): "
            ).strip().strip('"').strip("'")
        except Exception:
            user_input = ""
        if user_input:
            chosen = Path(user_input).expanduser()
            # If a file path was pasted, use its parent
            if chosen.is_file():
                chosen = chosen.parent
            if chosen.exists() and _has_client_lib(chosen):
                try:
                    oracledb.init_oracle_client(lib_dir=str(chosen))
                    used_lib_dir = str(chosen)
                except Exception as init_exc:
                    print(
                        f"Warning: failed to initialize Oracle Client at {str(chosen)!r}: {init_exc}",
                        file=sys.stderr,
                    )
            else:
                print(
                    f"Provided path does not look like an Instant Client directory: {str(chosen)!r}",
                    file=sys.stderr,
                )

    dsn = oracledb.makedsn(cfg["ORACLE_HOST"], int(cfg["ORACLE_PORT"]), sid=cfg["ORACLE_SID"])
    try:
        connection = oracledb.connect(
            user=cfg["ORACLE_USER"],
            password=cfg["ORACLE_PASSWORD"],
            dsn=dsn,
        )
    except (oracledb.NotSupportedError, oracledb.OperationalError) as exc:
        # If server mandates Native Network Encryption, thin mode cannot be used.
        if "DPY-3001" in str(exc) and used_lib_dir is None:
            raise RuntimeError(
                "Server requires Native Network Encryption, which needs python-oracledb thick mode.\n"
                "Install Oracle Instant Client for Windows (Basic or Basic Light) and set the client folder (with oci.dll):\n"
                "  - either in .env as ORACLE_CLIENT_LIB_DIR=C:\\path\\to\\instantclient_XX_X\n"
                "  - or as environment variable ORACLE_CLIENT_LIB_DIR, or add that folder to PATH.\n"
                "Then rerun this script."
            ) from exc
        raise

    connection.call_timeout = 120_000
    return connection


def float_matches(left: float | int | None, right: float | int) -> bool:
    if left is None:
        return False
    return abs(float(left) - float(right)) < 1e-9


def load_profile_sensors() -> list[ProfileSensor]:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["CO2"]

    sensors: list[ProfileSensor] = []
    for row_idx in range(4, sheet.max_row + 1):
        if sheet[f"C{row_idx}"].value != "C_CO2,basalt":
            continue
        if sheet[f"I{row_idx}"].value != "GMM222":
            continue
        if sheet[f"G{row_idx}"].value != SLOPE:
            continue
        if not float_matches(sheet[f"U{row_idx}"].value, X_COORD_M):
            continue
        if not float_matches(sheet[f"V{row_idx}"].value, Y_COORD_M):
            continue

        sensors.append(
            ProfileSensor(
                sensor_id=int(sheet[f"J{row_idx}"].value),
                sensor_code=str(sheet[f"K{row_idx}"].value),
                table_name=str(sheet[f"L{row_idx}"].value),
                variable_id=int(sheet[f"AG{row_idx}"].value),
                depth_m=float(sheet[f"W{row_idx}"].value),
            )
        )

    sensors.sort(key=lambda sensor: sensor.depth_m, reverse=True)
    if not sensors:
        raise ValueError(
            f"No basalt CO2 rows were found for slope={SLOPE!r}, X={X_COORD_M}, Y={Y_COORD_M}."
        )

    return sensors


def fetch_measurements(
    cursor: oracledb.Cursor,
    sensor: ProfileSensor,
    start_dt: datetime,
    end_dt: datetime | None,
) -> list[Measurement]:
    params: dict[str, object] = {
        "sensor_id": sensor.sensor_id,
        "variable_id": sensor.variable_id,
        "start_dt": start_dt,
        "no_data_value": NO_DATA_VALUE,
    }

    end_clause = ""
    if end_dt is not None:
        end_clause = "AND dv.localdatetime <= :end_dt"
        params["end_dt"] = end_dt

    query = f"""
        SELECT
            dv.localdatetime,
            dv.datavalue
        FROM
            {sensor.table_name} dv
        WHERE
            dv.sensorid = :sensor_id
            AND dv.variableid = :variable_id
            AND dv.localdatetime >= :start_dt
            {end_clause}
            AND dv.datavalue <> :no_data_value
        ORDER BY
            dv.localdatetime
    """

    cursor.execute(query, params)
    return [Measurement(timestamp=row[0], value=float(row[1])) for row in cursor.fetchall()]


def resolve_end_datetime(
    requested_end_dt: datetime | None,
    series_by_depth: dict[float, list[Measurement]],
) -> datetime:
    if requested_end_dt is not None:
        return requested_end_dt

    available_last_times = [
        series[-1].timestamp for series in series_by_depth.values() if series
    ]
    if not available_last_times:
        raise ValueError("No observations were found for the selected profile and date range.")
    return max(available_last_times)


def build_frames(
    series_by_depth: dict[float, list[Measurement]],
    end_dt: datetime,
) -> list[tuple[datetime, dict[float, float]]]:
    events_by_time: dict[datetime, list[tuple[float, float]]] = defaultdict(list)
    for depth_m, series in series_by_depth.items():
        for measurement in series:
            if measurement.timestamp <= end_dt:
                events_by_time[measurement.timestamp].append((depth_m, measurement.value))

    frame_times = sorted(events_by_time)
    if not frame_times:
        raise ValueError("No frame timestamps were built for the selected profile and date range.")

    current_values: dict[float, float] = {}
    frames: list[tuple[datetime, dict[float, float]]] = []
    for frame_time in frame_times:
        for depth_m, value in events_by_time[frame_time]:
            current_values[depth_m] = value
        frames.append((frame_time, current_values.copy()))

    return frames


def format_datetime_for_display(value: datetime) -> str:
    return value.strftime(DATETIME_WITH_TIME_FMT)


def format_coordinate(value: float | int) -> str:
    as_float = float(value)
    if as_float.is_integer():
        return str(int(as_float))
    return f"{as_float:g}"


def sanitize_filename_part(text: str) -> str:
    return text.replace(" ", "_")


def build_output_stem(start_dt: datetime, end_dt: datetime) -> str:
    slope_part = sanitize_filename_part(SLOPE)
    x_part = format_coordinate(X_COORD_M)
    y_part = format_coordinate(Y_COORD_M)
    start_part = start_dt.strftime("%Y-%b-%d_%H-%M")
    end_part = end_dt.strftime("%Y-%b-%d_%H-%M")
    return f"co2_profile_{slope_part}_x{x_part}_y{y_part}_{start_part}_to_{end_part}"


def draw_frame(
    figure: plt.Figure,
    axis: plt.Axes,
    frame_time: datetime,
    current_values: dict[float, float],
) -> None:
    axis.clear()

    axis.set_xlim(CO2_AXIS_MIN_PPM, CO2_AXIS_MAX_PPM)
    axis.set_ylim(-0.55, 0.03)
    axis.set_xlabel("CO2 concentration [ppm]")
    axis.set_ylabel("Depth [m]")
    axis.set_yticks(DISPLAY_DEPTHS_M)
    axis.set_yticklabels([f"{depth:.2f}" if depth != 0 else "0.00" for depth in DISPLAY_DEPTHS_M])
    axis.grid(axis="x", color=GRID_COLOR, linewidth=0.8, alpha=0.7)
    axis.set_axisbelow(True)
    axis.axhline(0.0, color=SURFACE_LINE_COLOR, linewidth=1.3)

    for depth_m in DISPLAY_DEPTHS_M[1:]:
        if depth_m not in current_values:
            continue

        value = current_values[depth_m]
        visible_width = max(value, 0.0)
        axis.barh(
            y=depth_m,
            width=visible_width,
            height=BAR_HEIGHT_M,
            left=0.0,
            color=BAR_COLOR,
            edgecolor="black",
            linewidth=0.8,
        )

        if value >= CO2_AXIS_MAX_PPM * 0.92:
            label_x = CO2_AXIS_MAX_PPM * 0.98
            ha = "right"
        else:
            label_x = max(visible_width + 90, CO2_AXIS_MIN_PPM + 50)
            ha = "left"
        axis.text(
            label_x,
            depth_m,
            f"{value:.0f} ppm",
            va="center",
            ha=ha,
            fontsize=10,
            clip_on=False,
        )

    values_text = []
    for depth_m in DISPLAY_DEPTHS_M[1:]:
        if depth_m in current_values:
            values_text.append(f"{depth_m:.2f} m = {current_values[depth_m]:.0f} ppm")
    values_block = "\n".join(values_text) if values_text else "No values available in this frame"

    axis.set_title(
        f"{SLOPE} | X={format_coordinate(X_COORD_M)} m | Y={format_coordinate(Y_COORD_M)} m | "
        f"{format_datetime_for_display(frame_time)}",
        fontsize=12,
        pad=12,
    )
    axis.text(
        0.98,
        0.97,
        values_block,
        transform=axis.transAxes,
        ha="right",
        va="top",
        fontsize=9.5,
        bbox={"facecolor": "white", "edgecolor": "#888888", "alpha": 0.92, "boxstyle": "round,pad=0.35"},
    )

    figure.tight_layout()


def main() -> None:
    start_dt = parse_user_datetime(START_DATE, is_end=False)
    if start_dt is None:
        raise ValueError("START_DATE must be provided.")
    requested_end_dt = parse_user_datetime(END_DATE, is_end=True)
    if requested_end_dt is not None and requested_end_dt < start_dt:
        raise ValueError("END_DATE must be greater than or equal to START_DATE.")

    profile_sensors = load_profile_sensors()

    connection = connect_to_oracle()
    cursor = connection.cursor()
    try:
        series_by_depth = {
            sensor.depth_m: fetch_measurements(cursor, sensor, start_dt, requested_end_dt)
            for sensor in profile_sensors
        }
    finally:
        cursor.close()
        connection.close()

    resolved_end_dt = resolve_end_datetime(requested_end_dt, series_by_depth)
    frames = build_frames(series_by_depth, resolved_end_dt)

    output_stem = build_output_stem(start_dt, resolved_end_dt)
    output_gif = OUTPUT_DIR / f"{output_stem}.gif"
    output_jpg = OUTPUT_DIR / f"{output_stem}.jpg"

    plt.rcParams.update(
        {
            "figure.figsize": (10.5, 6.5),
            "figure.dpi": 120,
            "font.size": 11,
            "axes.titlesize": 12,
            "axes.labelsize": 11,
        }
    )

    figure, axis = plt.subplots()

    def update(frame_index: int) -> list[object]:
        frame_time, current_values = frames[frame_index]
        draw_frame(figure, axis, frame_time, current_values)
        return []

    animation = FuncAnimation(
        figure,
        update,
        frames=len(frames),
        interval=FRAME_DURATION_MS,
        repeat=False,
        blit=False,
    )

    fps = max(1, round(1000 / FRAME_DURATION_MS))
    writer = SinglePlayPillowWriter(fps=fps)
    print(f"Saving GIF to {output_gif}")
    animation.save(
        output_gif,
        writer=writer,
        dpi=160,
        progress_callback=lambda current_frame, total_frames: (
            print(f"Frame {current_frame + 1}/{total_frames}", end="\r")
            if (current_frame + 1) % 10 == 0 or current_frame + 1 == total_frames
            else None
        ),
    )
    print()

    update(len(frames) - 1)
    print(f"Saving final JPEG to {output_jpg}")
    figure.savefig(output_jpg, format="jpeg", dpi=200)
    plt.close(figure)

    print(f"Done. Frames: {len(frames)}")
    print(f"GIF: {output_gif}")
    print(f"JPEG: {output_jpg}")


if __name__ == "__main__":
    main()
